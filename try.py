from openpyxl import load_workbook
from biblio import clientes
from win32com import client
from sys import exit
from PySimpleGUI import PySimpleGUI as sg
from datetime import datetime
import os

#Definir Data Padrão
data_atual = datetime.today()
data_texto = data_atual.strftime('%d/%m/%y')
dataPedido = data_texto
#Preços Padrão
precobb = 190
precogv = 175
precoverm = 205
precopto = 220
precoscbb = 365
precoscgv = 340
#Definir Layouts das Janelas
while True:
    def tela_pedido():
        sg.theme('Reddit')
        layout = [
            [sg.Text('Cód:'), sg.Input(key='codigo_cliente', size=(10, 1)), sg.Checkbox('S/N', key='nota'), sg.Button('Cadastrar',key='cadastro',size=(11,1))],
            [sg.Text('Data (Atual por padrão)'), sg.Button('Digitar Data',key='checkbox_data_nao',size=(11,1)), sg.Button('Limpar Data')],
            #[sg.Text('Preços (Atuais por padrão)'), sg.Button('Digitar Preços',key='checkbox_preco_nao'), sg.Button('Limpar Preços')],
            [sg.Text('Quantidades'),sg.Text('Preços'),sg.Text('R$')],
            [sg.Text('Barbalho 1kg'), sg.Input(f'{int(0)}',key='qtd_bb_1kg', size=(10, 1)),sg.Button('+1',key='+bb1'),sg.Button('-1',key='-bb1'),sg.Input(f'{int(precobb)}',key='prc_bb_1kg', size=(10, 1)),sg.Button('+1',key='+bb1prc'),sg.Button('-1',key='-bb1prc')],
            [sg.Text('Barbalho 2kg'), sg.Input(f'{int(0)}',key='qtd_bb_2kg', size=(10, 1)),sg.Button('+1',key='+bb2'),sg.Button('-1',key='-bb2'),sg.Input(f'{int(precobb)}',key='prc_bb_2kg', size=(10, 1)),sg.Button('+1',key='+bb2prc'),sg.Button('-1',key='-bb2prc')],
            [sg.Text('Barbalho 5kg'), sg.Input(f'{int(0)}',key='qtd_bb_5kg', size=(10, 1)),sg.Button('+1',key='+bb5'),sg.Button('-1',key='-bb5'),sg.Input(f'{int(precobb)}',key='prc_bb_5kg', size=(10, 1)),sg.Button('+1',key='+bb5prc'),sg.Button('-1',key='-bb5prc')],
            [sg.Text('Vermelho     '), sg.Input(f'{int(0)}',key='qtd_verm', size=(10, 1)),sg.Button('+1',key='+verm'),sg.Button('-1',key='-verm'),sg.Input(f'{int(precoverm)}',key='prc_verm', size=(10, 1)),sg.Button('+1',key='+vermprc'),sg.Button('-1',key='-vermprc')],
            [sg.Text('Preto           '), sg.Input(f'{int(0)}',key='qtd_pto', size=(10, 1)),sg.Button('+1',key='+pto'),sg.Button('-1',key='-pto'),sg.Input(f'{int(precopto)}',key='prc_pto', size=(10, 1)),sg.Button('+1',key='+ptoprc'),sg.Button('-1',key='-ptoprc')],
            [sg.Text('Goval 1kg    '), sg.Input(f'{int(0)}',key='qtd_gv_1kg', size=(10, 1)),sg.Button('+1',key='+gv1'),sg.Button('-1',key='-gv1'),sg.Input(f'{int(precogv)}',key='prc_gv_1kg', size=(10, 1)),sg.Button('+1',key='+gv1prc'),sg.Button('-1',key='-gv1prc')],
            [sg.Text('Goval 5kg    '), sg.Input(f'{int(0)}',key='qtd_gv_5kg', size=(10, 1)),sg.Button('+1',key='+gv5'),sg.Button('-1',key='-gv5'),sg.Input(f'{int(precogv)}',key='prc_gv_5kg', size=(10, 1)),sg.Button('+1',key='+gv5prc'),sg.Button('-1',key='-gv5prc')],
            [sg.Text('Sc. Barbalho'), sg.Input(f'{int(0)}',key='qtd_sc_bb', size=(10, 1)),sg.Button('+1',key='+scbb'),sg.Button('-1',key='-scbb'),sg.Input(f'{int(precoscbb)}',key='prc_sc_bb', size=(10, 1)),sg.Button('+1',key='+scbbprc'),sg.Button('-1',key='-scbbprc')],
            [sg.Text('Sc. Goval     '), sg.Input(f'{int(0)}',key='qtd_sc_gv', size=(10, 1)),sg.Button('+1',key='+scgv'),sg.Button('-1',key='-scgv'),sg.Input(f'{int(precoscgv)}',key='prc_sc_gv', size=(10, 1)),sg.Button('+1',key='+scgvprc'),sg.Button('-1',key='-scgvprc')],
            [sg.Text('Observações'), sg.Input(key='obs', size=(42, 1))],
            [sg.Button('Enviar'), sg.Button('Fechar')],
            #[sg.Output(size=(30,10))]
        ]
        return sg.Window('telaPedido',layout=layout,finalize=True)


    def precos():
        sg.theme('Reddit')
        layout = [
            [sg.Text('Usar Preços')],
            [sg.Text('Barbalho R$'),sg.Input(key='preco_bb')],
            [sg.Text('Goval R$'),sg.Input(key='preco_gv')],
            [sg.Text('Vermelho R$'),sg.Input(key='preco_verm')],
            [sg.Text('Preto R$'),sg.Input(key='preco_pto')],
            [sg.Text('Sc. Barbalho R$'),sg.Input(key='preco_sc_bb')],
            [sg.Text('Sc. Goval R$'),sg.Input(key='preco_sc_gv')],
            [sg.Button('OK'),sg.Button('VOLTAR')]
        ]
        return sg.Window('inserirPrecos',layout=layout,finalize=True)


    def janelaData():
        sg.theme('Reddit')
        layout = [
            [sg.Text('Data [dd/mm/aa]'),sg.Input(key='data_dia')],
            [sg.Button('OK'),sg.Button('Voltar')]
        ]
        return sg.Window('inserirData',layout=layout,finalize=True)


    def janelaCadastro():
        sg.theme('Reddit')
        layout = [
            [sg.Text('Cadastrar')],
            [sg.Text('Cliente'),sg.Input(key='clienteCadastro')],
            #[sg.Text('Endereço'), sg.Input(key='endereçoCadastro')],
            [sg.Text('Cidade'), sg.Input(key='cidadeCadastro')],
            #[sg.Text('Fone'), sg.Input(key='foneCadastro')],
            #[sg.Text('CNPJ'), sg.Input(key='cnpjCadastro')],
            #[sg.Text('Insc. Est.'), sg.Input(key='ieCadastro')],
            [sg.Text('Cond. de Pag.'), sg.Input(key='pagCadastro')],
            #[sg.Text('Email'), sg.Input(key='emailCadastro')],
            #[sg.Checkbox('S/N', key='nota')],
            [sg.Button('Enviar'), sg.Button('Fechar')],
        ]
        return sg.Window('janelaCadastro', layout=layout, finalize=True)


    def atualizarVisorSoma(valor,elemento):
        soma = int(valor) + int(1)
        try:
            janela1.Element(elemento).update(value=soma)
        except Exception as e:
            print(f'erro {e}')


    def atualizarVisorSub(valor,elemento):
        if int(valor) == 0:
            sub = 0
        else:
            sub = int(valor) - int(1)
        try:
            janela1.Element(elemento).update(value=sub)
        except Exception as e:
            print(f'erro {e}')


#Definindo Janelas (Janela1 = Janela inicial)
    janela1,janela2,janela3,janela4 = tela_pedido(), None, None,None


    #Ler os eventos
    while True:
        janela, evento, valores = sg.read_all_windows()
    #Eventos Janela1
        #Fechar Janela
        if janela == janela1 and evento == 'Fechar':
            exit()
        if janela == janela1 and evento == sg.WIN_CLOSED:
            exit()

        #Redefinir Data para Padrão
        if janela == janela1 and evento == 'Limpar Data':
            dataPedido = data_texto
            print(f'Data Definida: {dataPedido}')
            print()

        #Definir Função dos botões "+1" e "-1"
        if janela == janela1 and evento == '+bb1':
            atualizarVisorSoma(valores['qtd_bb_1kg'],'qtd_bb_1kg')
        if janela == janela1 and evento == '-bb1':
            atualizarVisorSub(valores['qtd_bb_1kg'],'qtd_bb_1kg')

        if janela == janela1 and evento == '+bb1prc':
            atualizarVisorSoma(valores['prc_bb_1kg'],'prc_bb_1kg')
        if janela == janela1 and evento == '-bb1prc':
            atualizarVisorSub(valores['prc_bb_1kg'],'prc_bb_1kg')

        if janela == janela1 and evento == '+bb2':
            atualizarVisorSoma(valores['qtd_bb_2kg'],'qtd_bb_2kg')
        if janela == janela1 and evento == '-bb2':
            atualizarVisorSub(valores['qtd_bb_2kg'],'qtd_bb_2kg')

        if janela == janela1 and evento == '+bb2prc':
            atualizarVisorSoma(valores['prc_bb_2kg'],'prc_bb_2kg')
        if janela == janela1 and evento == '-bb2prc':
            atualizarVisorSub(valores['prc_bb_2kg'],'prc_bb_2kg')

        if janela == janela1 and evento == '+bb5':
            atualizarVisorSoma(valores['qtd_bb_5kg'],'qtd_bb_5kg')
        if janela == janela1 and evento == '-bb5':
            atualizarVisorSub(valores['qtd_bb_5kg'],'qtd_bb_5kg')

        if janela == janela1 and evento == '+bb5prc':
            atualizarVisorSoma(valores['prc_bb_5kg'],'prc_bb_5kg')
        if janela == janela1 and evento == '-bb5prc':
            atualizarVisorSub(valores['prc_bb_5kg'],'prc_bb_5kg')

        if janela == janela1 and evento == '+verm':
            atualizarVisorSoma(valores['qtd_verm'],'qtd_verm')
        if janela == janela1 and evento == '-verm':
            atualizarVisorSub(valores['qtd_verm'],'qtd_verm')

        if janela == janela1 and evento == '+vermprc':
            atualizarVisorSoma(valores['prc_verm'],'prc_verm')
        if janela == janela1 and evento == '-vermprc':
            atualizarVisorSub(valores['prc_verm'],'prc_verm')

        if janela == janela1 and evento == '+pto':
            atualizarVisorSoma(valores['qtd_pto'],'qtd_pto')
        if janela == janela1 and evento == '-pto':
            atualizarVisorSub(valores['qtd_pto'],'qtd_pto')

        if janela == janela1 and evento == '+ptoprc':
            atualizarVisorSoma(valores['prc_pto'],'prc_pto')
        if janela == janela1 and evento == '-ptoprc':
            atualizarVisorSub(valores['prc_pto'],'prc_pto')

        if janela == janela1 and evento == '+gv1':
            atualizarVisorSoma(valores['qtd_gv_1kg'],'qtd_gv_1kg')
        if janela == janela1 and evento == '-gv1':
            atualizarVisorSub(valores['qtd_gv_1kg'],'qtd_gv_1kg')


        if janela == janela1 and evento == '+gv1prc':
            atualizarVisorSoma(valores['prc_gv_1kg'],'prc_gv_1kg')
        if janela == janela1 and evento == '-gv1prc':
            atualizarVisorSub(valores['prc_gv_1kg'],'prc_gv_1kg')

        if janela == janela1 and evento == '+gv5':
            atualizarVisorSoma(valores['qtd_gv_5kg'],'qtd_gv_5kg')
        if janela == janela1 and evento == '-gv5':
            atualizarVisorSub(valores['qtd_gv_5kg'],'qtd_gv_5kg')

        if janela == janela1 and evento == '+gv5prc':
            atualizarVisorSoma(valores['prc_gv_5kg'],'prc_gv_5kg')
        if janela == janela1 and evento == '-gv5prc':
            atualizarVisorSub(valores['prc_gv_5kg'],'prc_gv_5kg')

        if janela == janela1 and evento == '+scbb':
            atualizarVisorSoma(valores['qtd_sc_bb'],'qtd_sc_bb')
        if janela == janela1 and evento == '-scbb':
            atualizarVisorSub(valores['qtd_sc_bb'],'qtd_sc_bb')


        if janela == janela1 and evento == '+scbbprc':
            atualizarVisorSoma(valores['prc_sc_bb'],'prc_sc_bb')
        if janela == janela1 and evento == '-scbbprc':
            atualizarVisorSub(valores['prc_sc_bb'],'prc_sc_bb')

        if janela == janela1 and evento == '+scgv':
            atualizarVisorSoma(valores['qtd_sc_gv'],'qtd_sc_gv')
        if janela == janela1 and evento == '-scgv':
            atualizarVisorSub(valores['qtd_sc_gv'],'qtd_sc_gv')

        if janela == janela1 and evento == '+scgvprc':
            atualizarVisorSoma(valores['prc_sc_gv'],'prc_sc_gv')
        if janela == janela1 and evento == '-scgvprc':
            atualizarVisorSub(valores['prc_sc_gv'],'prc_sc_gv')

        #Abrir Janela de Data
        if janela == janela1 and evento == 'checkbox_data_nao':
            janela3 = janelaData()
            janela1.hide()
        #Definir Função do Botão Enviar
        if janela == janela1 and evento == 'Enviar':
            if valores['nota'] == True:
                notaPedido = 'S/N'
            elif valores['nota'] == False:
                notaPedido = ''
            try:
                codigoPedido = int(valores['codigo_cliente'])
                precobb1 = valores['prc_bb_1kg']
                precobb2 = valores['prc_bb_2kg']
                precobb5 = valores['prc_bb_5kg']
                precogv1 = valores['prc_gv_1kg']
                precogv5 = valores['prc_gv_5kg']
                precoverm = valores['prc_verm']
                precopto = valores['prc_pto']
                precoscbb = valores['prc_sc_bb']
                precoscgv = valores['prc_sc_gv']
                print(
                    f'Barbalho: {precobb}\nGoval: {precogv}\nVermelho: {precoverm}\nPreto: {precopto}\nSaco Barb.: {precoscbb}\nSaco Gov.: {precoscgv}')
                print()
                qtdBB1Pedido = int(valores['qtd_bb_1kg'])
                qtdBB2Pedido = int(valores['qtd_bb_2kg'])
                qtdBB5Pedido = int(valores['qtd_bb_5kg'])
                qtdVermPedido = int(valores['qtd_verm'])
                qtdPretoPedido = int(valores['qtd_pto'])
                qtdGV1Pedido = int(valores['qtd_gv_1kg'])
                qtdGV5Pedido = int(valores['qtd_gv_5kg'])
                qtdSCBBPedido = int(valores['qtd_sc_bb'])
                qtdSCGVPedido = int(valores['qtd_sc_gv'])
                obsPedido = valores['obs']
                janela1.hide()
                break
            except:
                print('Digite um código válido')

        #Função do Botão "Cadastro"aaaa
        if janela == janela1 and evento == 'cadastro':
            janela4 = janelaCadastro()
            janela1.hide()
    #Eventos Janela3
        # Função do Botão "OK"
        if janela == janela3 and evento == 'OK':
            janela3.hide()
            janela1.un_hide()
            dataPedido = valores['data_dia']
            print(f'Data Definida: {dataPedido}')
            print()
        # Função do Botão "Voltar"
        if janela == janela3 and evento == 'Voltar':
            janela3.hide()
            janela1.un_hide()
    #Eventos Janela4
        # Função do Botão "Fechar"
        if janela == janela4 and evento == 'Fechar':
            janela4.hide()
            janela1.un_hide()
        # Função do Botão "Enviar"
        if janela == janela4 and evento == 'Enviar':
            arquivo = 'cadastrosclientes.txt'
            try:
                clientes.cadastrar(arquivo, valores['clienteCadastro'], valores['cidadeCadastro'], valores['pagCadastro'])
                print('Cadastro Concluído')
            except Exception as e:
                print(f'Erro {e}')
            finally:
                janela4.hide()
                janela1.un_hide()

    #Abrir Bloco de Pedido no Excel
    diretorioexcel = os.getcwd()
    nomebloco = "BLOCOPROJETO.xlsx"
    print('CRIANDO ARQUIVO')
    arquivo = f"{diretorioexcel}\{nomebloco}"
    wb = load_workbook(arquivo)
    ws = wb.worksheets[0]

    #Programa Principal

    date = dataPedido
    codigo = codigoPedido
    cliente = clientes.itensArquivo(codigo)
    nome_cliente = cliente[0]
    cidade1 = cliente[1]
    pagamento = cliente[2]
    #Cabecalho
    ws['C2'] = f'{nome_cliente}'
    ws['C4'] = f'{cidade1}'
    ws['H6'] = f'{pagamento}'
    ws['Q2'] = f'{date}'
    ws['Q3'] = f'{notaPedido}'

    #Quantidades
    qtdbb1 = ws['A8'] = int(f'{qtdBB1Pedido}')
    qtdbb2 = ws['A9'] = int(f'{qtdBB2Pedido}')
    qtdbb5 = ws['A10'] = int(f'{qtdBB5Pedido}')
    qtdverm = ws['A11'] = int(f'{qtdVermPedido}')
    qtdpt = ws['A12'] = int(f'{qtdPretoPedido}')
    qtdgv1 = ws['A17'] = int(f'{qtdGV1Pedido}')
    qtdgv5 = ws['A18'] = int(f'{qtdGV5Pedido}')
    qtscbb = ws['A26'] = int(f'{qtdSCBBPedido}')
    qtscgv = ws['A27'] = int(f'{qtdSCGVPedido}')
    obser = ws['H33'] = f'{obsPedido}'
    #Preços Unitários
    if qtdBB1Pedido > 0:
        pubb1 = ws['O8'] = int(f'{precobb1}')
    if qtdBB2Pedido > 0:
        pubb2 = ws['O9'] = int(f'{precobb2}')
    if qtdBB5Pedido > 0:
        pubb5 = ws['O10'] = int(f'{precobb5}')
    if qtdVermPedido > 0:
        puv = ws['O11'] = int(f'{precoverm}')
    if qtdPretoPedido > 0:
        pup = ws['O12'] = int(f'{precopto}')
    if qtdGV1Pedido > 0:
        pugv1 = ws['O17'] = int(f'{precogv1}')
    if qtdGV5Pedido > 0:
        pugv5 = ws['O18'] = int(f'{precogv5}')
    if qtdSCBBPedido > 0:
        puscbb = ws['O26'] = int(f'{precoscbb}')
    if qtdSCGVPedido > 0:
        puscgv = ws['O27'] = int(f'{precoscgv}')

    #Salvar pedido XL
    diretorio = os.getcwd()

    pastaxl = 'PedidosXL'
    pastapdf = 'PedidosPDF'

    if os.path.isdir(pastaxl):  # vemos de este diretorio ja existe
        print('Ja existe uma pasta com esse nome!')
    else:
        os.mkdir(pastaxl)  # aqui criamos a pasta caso nao exista
        print('Pasta criada com sucesso!')

    if os.path.isdir(pastapdf):  # vemos se este diretorio ja existe
        print('Ja existe uma pasta com esse nome!')
    else:
        os.mkdir(pastapdf)  # aqui criamos a pasta caso nao exista
        print('Pasta criada com sucesso!')

    caminhoXL = f'{diretorio}\{pastaxl}'
    caminhoPDF = f'{diretorio}\{pastapdf}'

    nome_arquivo = fr'{caminhoXL}\Pedido_{nome_cliente.strip()}'

    wb.save(f'{nome_arquivo}.xlsx')
    #Criar Arquivo PDF

    nome_arquivo_PDF = fr'{caminhoPDF}\Pedido_{nome_cliente.strip()}'

    input_file = fr'{nome_arquivo}.xlsx'
    #give your file name with valid path
    output_file = fr'{nome_arquivo_PDF}.pdf'
    #give valid output file name and path
    app = client.DispatchEx("Excel.Application")
    app.Interactive = False
    app.Visible = False
    Workbook = app.Workbooks.Open(input_file)
    try:
        Workbook.ActiveSheet.ExportAsFixedFormat(0, output_file)
        print('Arquivo PDF Criado!')
    except Exception as e:
        print("Failed to convert in PDF format.Please confirm environment meets all the requirements  and try again")
        print(str(e))
    finally:
        try:
            Workbook.Close()
            app.Exit()
        except Exception as e:
            print(f'Erro {e}')