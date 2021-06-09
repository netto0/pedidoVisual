from openpyxl import load_workbook
from biblio import clientes
from win32com import client
from sys import exit
from PySimpleGUI import PySimpleGUI as sg
from datetime import datetime
import os
from biblio import janelas
from unidecode import unidecode
import re
import keyboard

# Definir Data Padrão
data_atual = datetime.today()
data_texto = data_atual.strftime('%d/%m/%y')
razoes = []


def give_razao():
    for c in range(0, 68):
        razoes.append(clientes.itensArquivo(c)[0])


give_razao()

# Preços Padrão
precobb = 200
precogv = 190
precoverm = 235
precopto = 210
precoscbb = 390
precoscgv = 370

# Tamanho Padrão
tam_horiz = 430
tam_vert = 773
tam_I = 57
tam_B = 24
tam_Qtd = (4,1)
tam_btt = (5,1)
# Definir Layouts das Janelas
while True:
    def tela_pedido():
        formaPagamento = ('Ch.', 'Bol.')
        prazoPagamento = (
            'À Vista', '14 dias', '28 dias', '30 dias', '35 dias', '40 dias', '45 dias', '50 dias', '28/35 dias',
            '30/40 dias', '30/45 dias', '30/60 dias', '35/45 dias', '30/40/50 dias')
        sg.theme('DarkBlue2')
        layout = [
            [sg.I(key='-SEARCH-', size=(tam_I, 1), enable_events=True)],
            [sg.LB(razoes, size=(tam_I, 15), key='-RAZOES-')],
            [sg.B('OK',size=(tam_B,1)), sg.B('Cadastrar',size=(tam_B,1))],
            [sg.T(f'Data:',size=(4,1),border_width=3), sg.I(size=(8,1), key='-DATA-', default_text=data_texto,text_color='yellow' ),
             sg.B('Hoje',size=(9,1)),sg.Checkbox('S/N', key='nota',),sg.T(f'Forma:',size=(8,1),justification='r'), sg.Combo(formaPagamento,key='-FORMA-', size=(5, 1),text_color='yellow')],
            [sg.T(f'Razão:', key='razao', size=(5,1)),sg.T(f'', key='razaoPedido', size=(24, 1),font=('Helvetica', 10),text_color='yellow'),sg.T('  Prazo:'),sg.Combo(prazoPagamento, key=('-PRAZO-'),size=(9,1),text_color='yellow')],
            [sg.T('Produto',size=(10,1),font=('Helvetica', 10,'bold'),justification='c'), sg.T('Qtd.',size=(9,1),border_width=0,font=('Helvetica', 10,'bold'),justification='l'), sg.T('R$',size=(12,1),font=('Helvetica', 10,'bold'),justification='r')],
            [sg.T('Barbalho 1kg',key='barbalho1kg',size=(10,1)), sg.I(f'{int(0)}', key='qtd_bb_1kg', size=tam_Qtd),sg.B('+1', key='+bb1',size=tam_btt),
             sg.B('-1', key='-bb1',size=tam_btt), sg.I(f'{int(precobb)}', key='prc_bb_1kg',size=tam_Qtd,justification='c'),
             sg.B('+1', key='+bb1prc',size=tam_btt), sg.B('-1', key='-bb1prc',size=tam_btt)],
            [sg.T('Barbalho 2kg',size=(10,1)), sg.I(f'{int(0)}', key='qtd_bb_2kg',size=tam_Qtd),
             sg.B('+1', key='+bb2',size=tam_btt),
             sg.B('-1', key='-bb2',size=tam_btt), sg.I(f'{int(precobb)}', key='prc_bb_2kg',size=tam_Qtd,justification='c'),
             sg.B('+1', key='+bb2prc',size=tam_btt), sg.B('-1', key='-bb2prc',size=tam_btt)],
            [sg.T('Barbalho 5kg',size=(10,1)), sg.I(f'{int(0)}', key='qtd_bb_5kg',size=tam_Qtd),
             sg.B('+1', key='+bb5',size=tam_btt),
             sg.B('-1', key='-bb5',size=tam_btt), sg.I(f'{int(precobb)}', key='prc_bb_5kg',size=tam_Qtd,justification='c'),
             sg.B('+1', key='+bb5prc',size=tam_btt), sg.B('-1', key='-bb5prc',size=tam_btt)],
            [sg.T('Vermelho',size=(10,1)), sg.I(f'{int(0)}', key='qtd_verm',size=tam_Qtd),
             sg.B('+1', key='+verm',size=tam_btt),
             sg.B('-1', key='-verm',size=tam_btt), sg.I(f'{int(precoverm)}', key='prc_verm',size=tam_Qtd,justification='c'),
             sg.B('+1', key='+vermprc',size=tam_btt), sg.B('-1', key='-vermprc',size=tam_btt)],
            [sg.T('Preto',size=(10,1)), sg.I(f'{int(0)}', key='qtd_pto',size=tam_Qtd),
             sg.B('+1', key='+pto',size=tam_btt),
             sg.B('-1', key='-pto',size=tam_btt), sg.I(f'{int(precopto)}', key='prc_pto',size=tam_Qtd,justification='c'),
             sg.B('+1', key='+ptoprc',size=tam_btt), sg.B('-1', key='-ptoprc',size=tam_btt)],
            [sg.T('Goval 1kg',size=(10,1)), sg.I(f'{int(0)}', key='qtd_gv_1kg',size=tam_Qtd),
             sg.B('+1', key='+gv1',size=tam_btt),
             sg.B('-1', key='-gv1',size=tam_btt), sg.I(f'{int(precogv)}', key='prc_gv_1kg',size=tam_Qtd,justification='c'),
             sg.B('+1', key='+gv1prc',size=tam_btt), sg.B('-1', key='-gv1prc',size=tam_btt)],
            [sg.T('Goval 5kg',size=(10,1)), sg.I(f'{int(0)}', key='qtd_gv_5kg',size=tam_Qtd),
             sg.B('+1', key='+gv5',size=tam_btt),
             sg.B('-1', key='-gv5',size=tam_btt), sg.I(f'{int(precogv)}', key='prc_gv_5kg',size=tam_Qtd,justification='c'),
             sg.B('+1', key='+gv5prc',size=tam_btt), sg.B('-1', key='-gv5prc',size=tam_btt)],
            [sg.T('Sc. Barbalho',size=(10,1)), sg.I(f'{int(0)}', key='qtd_sc_bb',size=tam_Qtd),
             sg.B('+1', key='+scbb',size=tam_btt),
             sg.B('-1', key='-scbb',size=tam_btt), sg.I(f'{int(precoscbb)}', key='prc_sc_bb',size=tam_Qtd,justification='c'),
             sg.B('+1', key='+scbbprc',size=tam_btt), sg.B('-1', key='-scbbprc',size=tam_btt)],
            [sg.T('Sc. Goval',size=(10,1)), sg.I(f'{int(0)}', key='qtd_sc_gv',size=tam_Qtd),
             sg.B('+1', key='+scgv',size=tam_btt),
             sg.B('-1', key='-scgv',size=tam_btt), sg.I(f'{int(precoscgv)}', key='prc_sc_gv',size=tam_Qtd,justification='c'),
             sg.B('+1', key='+scgvprc',size=tam_btt), sg.B('-1', key='-scgvprc',size=tam_btt)],
            [sg.T('Observações',size=(10,1)), sg.I(key='obs', size=(44, 1))],
            [sg.B('Enviar',size=(tam_B,1)), sg.B('Fechar',size=(tam_B,1))],
            # [sg.O(size=(70,10))]
        ]

        return sg.Window('telaPedido', layout=layout, finalize=True, return_keyboard_events=True,
                         size=(tam_horiz,tam_vert),margins=(10,10),border_depth=1)


    def precos():
        sg.theme('Reddit')
        layout = [
            [sg.Text('Usar Preços')],
            [sg.Text('Barbalho R$'), sg.Input(key='preco_bb')],
            [sg.Text('Goval R$'), sg.Input(key='preco_gv')],
            [sg.Text('Vermelho R$'), sg.Input(key='preco_verm')],
            [sg.Text('Preto R$'), sg.Input(key='preco_pto')],
            [sg.Text('Sc. Barbalho R$'), sg.Input(key='preco_sc_bb')],
            [sg.Text('Sc. Goval R$'), sg.Input(key='preco_sc_gv')],
            [sg.Button('OK'), sg.Button('VOLTAR')]
        ]
        return sg.Window('inserirPrecos', layout=layout, finalize=True)


    def atualizarVisorSoma(valor, elemento):
        soma = int(valor) + int(1)
        try:
            janela1.Element(elemento).update(value=soma)
        except Exception as e:
            print(f'erro {e}')


    def atualizarVisorSub(valor, elemento):
        if int(valor) == 0:
            sub = 0
        else:
            sub = int(valor) - int(1)
        try:
            janela1.Element(elemento).update(value=sub)
        except Exception as e:
            print(f'erro {e}')


    def to_ascii(ls):
        for i in range(len(ls)):
            ls[i] = unidecode(ls[i])


    nova_lista = []


    def search(nome, lista, indice=False):
        nova_lista = []
        to_ascii(lista)
        ref = unidecode(nome)
        for l in lista:
            if re.findall(rf'{ref}', l, flags=re.I) != []:
                numero = lista.index(l)
                nova_lista.append(razoes[numero])
            else:
                var = None
        if indice:
            return lista.index(nova_lista[0])
        else:
            return nova_lista


    def kb_event_update():
        try:
            if janela == janela1 and len(evento) == 1 or keyboard.is_pressed('\b'):
                dta = search(valores['-SEARCH-'], razoes)
                janela1.Element('-RAZOES-').update(values=dta)
        except:
            var = None


    def data_format():
        data = valores['-DATA-']
        try:
            if janela == janela1 and len(evento) == 1 or keyboard.is_pressed('\b'):
                tamanho = len(data)
                if tamanho == 2 or tamanho == 5:
                    janela1.Element('-DATA-').update(value=f'{data}/')
                elif tamanho == 6 and str(data).isnumeric():
                    janela1.Element('-DATA-').update(value=f'{data[0:2]}/{data[2:4]}/{data[4:8]}')
        except:
            var = None


    # Definindo Janelas (Janela1 = Janela inicial)
    janela1, janela2, janela3, janela4 = tela_pedido(), None, None, None

    # Ler os eventos
    while True:
        janela, evento, valores = sg.read_all_windows()
        # Eventos Janela1
        kb_event_update()
        data_format()
        data_valida = False
        # Fechar Janela
        if janela == janela1 and evento == 'Fechar':
            exit()
        if janela == janela1 and evento == sg.WIN_CLOSED:
            exit()
        if janela == janela1 and evento == 'Cadastrar':
            janela1.hide()
            janela2 = janelas.janelaCadastro()
        if janela == janela1 and evento == 'Hoje':
            janela1.Element('-DATA-').update(value=data_texto)
        if janela == janela1 and evento == 'OK':
            try:
                if valores['-RAZOES-']:  # if something is highlighted in the list
                    codigoPedidoPrevia = search(valores['-RAZOES-'][0], razoes, indice=True)
                    clientePrevia = clientes.itensArquivo(codigoPedidoPrevia)
                    razao = clientePrevia[0]
                    formaPrevia = clientePrevia[2]
                    prazoPrevia = clientePrevia[3]
                    janela1.Element('razaoPedido').update(value=f'{razao}')
                    janela1.Element('-PRAZO-').update(value=f'{prazoPrevia} dias')
                    janela1.Element('-FORMA-').update(value=f'{formaPrevia}')
            except Exception as e:
                print('Digite um código válido')
                print(e)

        if janela == janela1 and evento == 'Clear':
            nova_lista = []
            janela1.Element('-RAZOES-').Update(values=razoes)
            janela1.Element('-SEARCH-').Update(value='')
            janela1.Element('razaoPedido').Update(value='')
            janela1.Element('-FORMA-').Update(value='')
            janela1.Element('-PRAZO-').Update(value='')

        # Definir Função dos botões "+1" e "-1"


        def buttons(evento1, evento2, valor):
            pevento1 = f'{evento1}'
            pevento2 = f'{evento2}'
            valorV = f'{valor}'

            if janela == janela1 and evento == pevento1:
                atualizarVisorSoma(valores[valorV], valorV)
            if janela == janela1 and evento == pevento2:
                atualizarVisorSub(valores[valorV], valorV)


        buttons('+bb1', '-bb1', 'qtd_bb_1kg')
        buttons('+bb1prc', '-bb1prc', 'prc_bb_1kg')
        buttons('+bb2', '-bb2', 'qtd_bb_2kg')
        buttons('+bb2prc', '-bb2prc', 'prc_bb_2kg')
        buttons('+bb5', '-bb5', 'qtd_bb_5kg')
        buttons('+bb5prc', '-bb5prc', 'prc_bb_5kg')
        buttons('+verm', '-verm', 'qtd_verm')
        buttons('+vermprc', '-vermprc', 'prc_verm')
        buttons('+pto', '-pto', 'qtd_pto')
        buttons('+ptoprc', '-ptoprc', 'prc_pto')
        buttons('+gv1', '-gv1', 'qtd_gv_1kg')
        buttons('+gv1prc', '-gv1prc', 'prc_gv_1kg')
        buttons('+gv5', '-gv5', 'qtd_gv_5kg')
        buttons('+gv5prc', '-gv5prc', 'prc_gv_5kg')
        buttons('+scbb', '-scbb', 'qtd_sc_bb')
        buttons('+scbbprc', '-scbbprc', 'prc_sc_bb')
        buttons('+scgv', '-scgv', 'qtd_sc_gv')
        buttons('+scgvprc', '-scgvprc', 'prc_sc_gv')

        if janela == janela1 and evento == 'Ver Lista':
            clientes.linhasArquivo('cadastrosclientes.txt')

        if janela == janela1 and evento == 'Enviar':  # Definir Função do Botão Enviar

            try:
                precobb1 = valores['prc_bb_1kg']
                precobb2 = valores['prc_bb_2kg']
                precobb5 = valores['prc_bb_5kg']
                precogv1 = valores['prc_gv_1kg']
                precogv5 = valores['prc_gv_5kg']
                precoverm = valores['prc_verm']
                precopto = valores['prc_pto']
                precoscbb = valores['prc_sc_bb']
                precoscgv = valores['prc_sc_gv']
                qtdBB1Pedido = int(valores['qtd_bb_1kg'])
                qtdBB2Pedido = int(valores['qtd_bb_2kg'])
                qtdBB5Pedido = int(valores['qtd_bb_5kg'])
                qtdVermPedido = int(valores['qtd_verm'])
                qtdPretoPedido = int(valores['qtd_pto'])
                qtdGV1Pedido = int(valores['qtd_gv_1kg'])
                qtdGV5Pedido = int(valores['qtd_gv_5kg'])
                qtdSCBBPedido = int(valores['qtd_sc_bb'])
                qtdSCGVPedido = int(valores['qtd_sc_gv'])
                obsPedido = valores['obs']
                date = valores['-DATA-']
                cliente = clientes.itensArquivo(codigoPedidoPrevia)
                nome_cliente = cliente[0]
                cidade1 = cliente[1]

                forma = valores['-FORMA-']
                prazo = valores['-PRAZO-']
                pagamento = f'{forma} {prazo}'

                if valores['nota']:
                    try:
                        forma.index('Ch')
                        notaPedido = 'S/N'
                        janela1.hide()
                        break
                    except:
                        print('Não é permitido enviar boleto sem nota')
                if not valores['nota']:
                    notaPedido = ''
                    janela1.hide()
                    break
                if len(valores["-DATA-"]) != 8:
                    print('Digite uma data válida!')
            except Exception as e:
                print('Digite um código válido')
                print(TypeError, e)

    # Eventos Janela4
    # Abrir Bloco de Pedido no Excel
    diretorioexcel = os.getcwd()
    nomebloco = "BLOCOPROJETO.xlsx"
    print('CRIANDO ARQUIVO')
    arquivo = f"{diretorioexcel}\{nomebloco}"
    wb = load_workbook(arquivo)
    ws = wb.worksheets[0]

    # Programa Principal

    # Cabecalho
    ws['C2'] = f'{nome_cliente}'
    ws['C4'] = f'{cidade1}'
    ws['H6'] = f'{pagamento}'
    ws['Q2'] = f'{date}'
    ws['Q3'] = f'{notaPedido}'

    # Quantidades
    qtdbb1 = ws['A8'] = int(f'{qtdBB1Pedido}')
    qtdbb2 = ws['A9'] = int(f'{qtdBB2Pedido}')
    qtdbb5 = ws['A10'] = int(f'{qtdBB5Pedido}')
    qtdverm = ws['A11'] = int(f'{qtdVermPedido}')
    qtdpt = ws['A12'] = int(f'{qtdPretoPedido}')
    qtdgv1 = ws['A17'] = int(f'{qtdGV1Pedido}')
    qtdgv5 = ws['A18'] = int(f'{qtdGV5Pedido}')
    qtscbb = ws['A26'] = int(f'{qtdSCBBPedido}')
    qtscgv = ws['A27'] = int(f'{qtdSCGVPedido}')
    obser = ws['H33'] = f'{obsPedido}'

    # Preços Unitários
    if qtdBB1Pedido > 0:
        pubb1 = ws['O8'] = int(f'{precobb1}')
    if qtdBB2Pedido > 0:
        pubb2 = ws['O9'] = int(f'{precobb2}')
    if qtdBB5Pedido > 0:
        pubb5 = ws['O10'] = int(f'{precobb5}')
    if qtdVermPedido > 0:
        puv = ws['O11'] = int(f'{precoverm}')
    if qtdPretoPedido > 0:
        pup = ws['O12'] = int(f'{precopto}')
    if qtdGV1Pedido > 0:
        pugv1 = ws['O17'] = int(f'{precogv1}')
    if qtdGV5Pedido > 0:
        pugv5 = ws['O18'] = int(f'{precogv5}')
    if qtdSCBBPedido > 0:
        puscbb = ws['O26'] = int(f'{precoscbb}')
    if qtdSCGVPedido > 0:
        puscgv = ws['O27'] = int(f'{precoscgv}')

    # Salvar pedido XL
    diretorio = os.getcwd()

    pastaxl = 'PedidosXL'
    pastapdf = 'PedidosPDF'

    if os.path.isdir(pastaxl):  # vemos se este diretório ja existe
        None
    else:
        os.mkdir(pastaxl)  # aqui criamos a pasta caso nao exista
        print('Pasta criada com sucesso!')

    if os.path.isdir(pastapdf):  # vemos se este diretório ja existe
        None
    else:
        os.mkdir(pastapdf)  # aqui criamos a pasta caso nao exista
        print('Pasta criada com sucesso!')

    caminhoXL = f'{diretorio}\{pastaxl}'
    caminhoPDF = f'{diretorio}\{pastapdf}'

    nome_arquivo = fr'{caminhoXL}\Pedido_{nome_cliente.strip()}'

    try:
        wb.save(f'{nome_arquivo}.xlsx')
        print('ARQUIVO XL CRIADO')
        nome_arquivo_PDF = fr'{caminhoPDF}\Pedido_{nome_cliente.strip()}'

        input_file = fr'{nome_arquivo}.xlsx'
        # give your file name with valid path
        output_file = fr'{nome_arquivo_PDF}.pdf'
        # give valid output file name and path
        app = client.DispatchEx("Excel.Application")
        app.Interactive = False
        app.Visible = False
        Workbook = app.Workbooks.Open(input_file)
        try:
            Workbook.ActiveSheet.ExportAsFixedFormat(0, output_file)
            print('Arquivo PDF Criado!')
        except Exception as e:
            print(
                "Failed to convert in PDF format.Please confirm environment meets all the requirements  and try again")
            print(str(e))
        finally:
            try:
                Workbook.Close()
                app.Exit()
            except Exception as e:
                None

    except Exception as e:
        print(f'Já existe um arquivo com esse nome')
