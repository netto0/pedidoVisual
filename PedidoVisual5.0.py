from openpyxl import load_workbook
from biblio import clientes
from win32com import client
from sys import exit
from PySimpleGUI import PySimpleGUI as sg
from datetime import datetime
import os
from biblio import janelas
#Definir Data Padrão
data_atual = datetime.today()
data_texto = data_atual.strftime('%d/%m/%y')
dataPedido = data_texto
#Preços Padrão
precobb = 190
precogv = 175
precoverm = 205
precopto = 220
precoscbb = 365
precoscgv = 340
#Definir Layouts das Janelas
while True:
    def tela_pedido():
        sg.theme('Reddit')
        layout = [
            [sg.Text('Cód:'), sg.Input(key='codigo_cliente', size=(6, 1)), sg.Button('OK'), sg.B('Ver Lista'), sg.Button('Cadastrar')],
            [sg.Button('Digitar Data',key='checkbox_data_nao',size=(11,1)), sg.Button('Limpar Data')],
            [sg.Text(f'Razão:',key='razaoPedido',size=(40,1)),sg.Text(f'Data: {data_texto}',key='dataPrevia')],
            [sg.Text(f'Cond. de Pag.:',key='pagPedido',size=(30,1))],
            [sg.Checkbox('S/N', key='nota'), sg.Text('Quantidades'),sg.Text('Preços'),sg.Text('R$')],
            [sg.Text('Barbalho 1kg'), sg.Input(f'{int(0)}',key='qtd_bb_1kg', size=(10, 1)),sg.Button('+1',key='+bb1'),sg.Button('-1',key='-bb1'),sg.Input(f'{int(precobb)}',key='prc_bb_1kg', size=(10, 1)),sg.Button('+1',key='+bb1prc'),sg.Button('-1',key='-bb1prc')],
            [sg.Text('Barbalho 2kg'), sg.Input(f'{int(0)}',key='qtd_bb_2kg', size=(10, 1)),sg.Button('+1',key='+bb2'),sg.Button('-1',key='-bb2'),sg.Input(f'{int(precobb)}',key='prc_bb_2kg', size=(10, 1)),sg.Button('+1',key='+bb2prc'),sg.Button('-1',key='-bb2prc')],
            [sg.Text('Barbalho 5kg'), sg.Input(f'{int(0)}',key='qtd_bb_5kg', size=(10, 1)),sg.Button('+1',key='+bb5'),sg.Button('-1',key='-bb5'),sg.Input(f'{int(precobb)}',key='prc_bb_5kg', size=(10, 1)),sg.Button('+1',key='+bb5prc'),sg.Button('-1',key='-bb5prc')],
            [sg.Text('Vermelho     '), sg.Input(f'{int(0)}',key='qtd_verm', size=(10, 1)),sg.Button('+1',key='+verm'),sg.Button('-1',key='-verm'),sg.Input(f'{int(precoverm)}',key='prc_verm', size=(10, 1)),sg.Button('+1',key='+vermprc'),sg.Button('-1',key='-vermprc')],
            [sg.Text('Preto           '), sg.Input(f'{int(0)}',key='qtd_pto', size=(10, 1)),sg.Button('+1',key='+pto'),sg.Button('-1',key='-pto'),sg.Input(f'{int(precopto)}',key='prc_pto', size=(10, 1)),sg.Button('+1',key='+ptoprc'),sg.Button('-1',key='-ptoprc')],
            [sg.Text('Goval 1kg    '), sg.Input(f'{int(0)}',key='qtd_gv_1kg', size=(10, 1)),sg.Button('+1',key='+gv1'),sg.Button('-1',key='-gv1'),sg.Input(f'{int(precogv)}',key='prc_gv_1kg', size=(10, 1)),sg.Button('+1',key='+gv1prc'),sg.Button('-1',key='-gv1prc')],
            [sg.Text('Goval 5kg    '), sg.Input(f'{int(0)}',key='qtd_gv_5kg', size=(10, 1)),sg.Button('+1',key='+gv5'),sg.Button('-1',key='-gv5'),sg.Input(f'{int(precogv)}',key='prc_gv_5kg', size=(10, 1)),sg.Button('+1',key='+gv5prc'),sg.Button('-1',key='-gv5prc')],
            [sg.Text('Sc. Barbalho'), sg.Input(f'{int(0)}',key='qtd_sc_bb', size=(10, 1)),sg.Button('+1',key='+scbb'),sg.Button('-1',key='-scbb'),sg.Input(f'{int(precoscbb)}',key='prc_sc_bb', size=(10, 1)),sg.Button('+1',key='+scbbprc'),sg.Button('-1',key='-scbbprc')],
            [sg.Text('Sc. Goval     '), sg.Input(f'{int(0)}',key='qtd_sc_gv', size=(10, 1)),sg.Button('+1',key='+scgv'),sg.Button('-1',key='-scgv'),sg.Input(f'{int(precoscgv)}',key='prc_sc_gv', size=(10, 1)),sg.Button('+1',key='+scgvprc'),sg.Button('-1',key='-scgvprc')],
            [sg.Text('Observações'), sg.Input(key='obs', size=(42, 1))],
            [sg.Button('Enviar'), sg.Button('Fechar')],
            #[sg.Output(size=(70,10))]
        ]
        return sg.Window('telaPedido',layout=layout,finalize=True)


    def precos():
        sg.theme('Reddit')
        layout = [
            [sg.Text('Usar Preços')],
            [sg.Text('Barbalho R$'),sg.Input(key='preco_bb')],
            [sg.Text('Goval R$'),sg.Input(key='preco_gv')],
            [sg.Text('Vermelho R$'),sg.Input(key='preco_verm')],
            [sg.Text('Preto R$'),sg.Input(key='preco_pto')],
            [sg.Text('Sc. Barbalho R$'),sg.Input(key='preco_sc_bb')],
            [sg.Text('Sc. Goval R$'),sg.Input(key='preco_sc_gv')],
            [sg.Button('OK'),sg.Button('VOLTAR')]
        ]
        return sg.Window('inserirPrecos',layout=layout,finalize=True)


    def janelaData():
        sg.theme('Reddit')
        layout = [
            [sg.Text('Data [dd/mm/aa]'),sg.Input(key='data_dia')],
            [sg.Button('OK'),sg.Button('Voltar')]
        ]
        return sg.Window('inserirData',layout=layout,finalize=True)


    def atualizarVisorSoma(valor,elemento):
        soma = int(valor) + int(1)
        try:
            janela1.Element(elemento).update(value=soma)
        except Exception as e:
            print(f'erro {e}')


    def atualizarVisorSub(valor,elemento):
        if int(valor) == 0:
            sub = 0
        else:
            sub = int(valor) - int(1)
        try:
            janela1.Element(elemento).update(value=sub)
        except Exception as e:
            print(f'erro {e}')


#Definindo Janelas (Janela1 = Janela inicial)
    janela1,janela2,janela3,janela4 = tela_pedido(), None, None,None


    #Ler os eventos
    while True:
        janela, evento, valores = sg.read_all_windows()
    #Eventos Janela1
        #Fechar Janela
        if janela == janela1 and evento == 'Fechar':
            exit()
        if janela == janela1 and evento == sg.WIN_CLOSED:
            exit()

        if janela == janela1 and evento == 'OK':
            try:
                codigoPedidoPrevia = int(valores['codigo_cliente'])
                clientePrevia = clientes.itensArquivo(codigoPedidoPrevia)
                razao = clientePrevia[0]
                pagamentoPrevia = clientePrevia[2]
                janela1.Element('razaoPedido').update(value=f'Razão: {razao}')
                janela1.Element('pagPedido').update(value=f'Cond. de Pag.: {pagamentoPrevia}')
            except Exception as e:
                print('Digite um código válido')
                print(e)

        #Redefinir Data para Padrão
        if janela == janela1 and evento == 'Limpar Data':
            dataPedido = data_texto
            print(f'Data Definida: {dataPedido}')
            print()

        if janela == janela1 and evento == 'Cadastrar':
            janela4 = janelas.janelaCadastro()

        #Definir Função dos botões "+1" e "-1"
        if janela == janela1 and evento == '+bb1':
            atualizarVisorSoma(valores['qtd_bb_1kg'],'qtd_bb_1kg')
        if janela == janela1 and evento == '-bb1':
            atualizarVisorSub(valores['qtd_bb_1kg'],'qtd_bb_1kg')

        if janela == janela1 and evento == '+bb1prc':
            atualizarVisorSoma(valores['prc_bb_1kg'],'prc_bb_1kg')
        if janela == janela1 and evento == '-bb1prc':
            atualizarVisorSub(valores['prc_bb_1kg'],'prc_bb_1kg')

        if janela == janela1 and evento == '+bb2':
            atualizarVisorSoma(valores['qtd_bb_2kg'],'qtd_bb_2kg')
        if janela == janela1 and evento == '-bb2':
            atualizarVisorSub(valores['qtd_bb_2kg'],'qtd_bb_2kg')

        if janela == janela1 and evento == '+bb2prc':
            atualizarVisorSoma(valores['prc_bb_2kg'],'prc_bb_2kg')
        if janela == janela1 and evento == '-bb2prc':
            atualizarVisorSub(valores['prc_bb_2kg'],'prc_bb_2kg')

        if janela == janela1 and evento == '+bb5':
            atualizarVisorSoma(valores['qtd_bb_5kg'],'qtd_bb_5kg')
        if janela == janela1 and evento == '-bb5':
            atualizarVisorSub(valores['qtd_bb_5kg'],'qtd_bb_5kg')

        if janela == janela1 and evento == '+bb5prc':
            atualizarVisorSoma(valores['prc_bb_5kg'],'prc_bb_5kg')
        if janela == janela1 and evento == '-bb5prc':
            atualizarVisorSub(valores['prc_bb_5kg'],'prc_bb_5kg')

        if janela == janela1 and evento == '+verm':
            atualizarVisorSoma(valores['qtd_verm'],'qtd_verm')
        if janela == janela1 and evento == '-verm':
            atualizarVisorSub(valores['qtd_verm'],'qtd_verm')

        if janela == janela1 and evento == '+vermprc':
            atualizarVisorSoma(valores['prc_verm'],'prc_verm')
        if janela == janela1 and evento == '-vermprc':
            atualizarVisorSub(valores['prc_verm'],'prc_verm')

        if janela == janela1 and evento == '+pto':
            atualizarVisorSoma(valores['qtd_pto'],'qtd_pto')
        if janela == janela1 and evento == '-pto':
            atualizarVisorSub(valores['qtd_pto'],'qtd_pto')

        if janela == janela1 and evento == '+ptoprc':
            atualizarVisorSoma(valores['prc_pto'],'prc_pto')
        if janela == janela1 and evento == '-ptoprc':
            atualizarVisorSub(valores['prc_pto'],'prc_pto')

        if janela == janela1 and evento == '+gv1':
            atualizarVisorSoma(valores['qtd_gv_1kg'],'qtd_gv_1kg')
        if janela == janela1 and evento == '-gv1':
            atualizarVisorSub(valores['qtd_gv_1kg'],'qtd_gv_1kg')

        if janela == janela1 and evento == '+gv1prc':
            atualizarVisorSoma(valores['prc_gv_1kg'],'prc_gv_1kg')
        if janela == janela1 and evento == '-gv1prc':
            atualizarVisorSub(valores['prc_gv_1kg'],'prc_gv_1kg')

        if janela == janela1 and evento == '+gv5':
            atualizarVisorSoma(valores['qtd_gv_5kg'],'qtd_gv_5kg')
        if janela == janela1 and evento == '-gv5':
            atualizarVisorSub(valores['qtd_gv_5kg'],'qtd_gv_5kg')

        if janela == janela1 and evento == '+gv5prc':
            atualizarVisorSoma(valores['prc_gv_5kg'],'prc_gv_5kg')
        if janela == janela1 and evento == '-gv5prc':
            atualizarVisorSub(valores['prc_gv_5kg'],'prc_gv_5kg')

        if janela == janela1 and evento == '+scbb':
            atualizarVisorSoma(valores['qtd_sc_bb'],'qtd_sc_bb')
        if janela == janela1 and evento == '-scbb':
            atualizarVisorSub(valores['qtd_sc_bb'],'qtd_sc_bb')

        if janela == janela1 and evento == '+scbbprc':
            atualizarVisorSoma(valores['prc_sc_bb'],'prc_sc_bb')
        if janela == janela1 and evento == '-scbbprc':
            atualizarVisorSub(valores['prc_sc_bb'],'prc_sc_bb')

        if janela == janela1 and evento == '+scgv':
            atualizarVisorSoma(valores['qtd_sc_gv'],'qtd_sc_gv')
        if janela == janela1 and evento == '-scgv':
            atualizarVisorSub(valores['qtd_sc_gv'],'qtd_sc_gv')

        if janela == janela1 and evento == '+scgvprc':
            atualizarVisorSoma(valores['prc_sc_gv'],'prc_sc_gv')
        if janela == janela1 and evento == '-scgvprc':
            atualizarVisorSub(valores['prc_sc_gv'],'prc_sc_gv')

        if janela == janela1 and evento == 'Ver Lista':
            clientes.linhasArquivo('cadastrosclientes.txt')

        #Abrir Janela de Data
        if janela == janela1 and evento == 'checkbox_data_nao':
            janela3 = janelaData()
            janela1.hide()
        #Definir Função do Botão Enviar
        if janela == janela1 and evento == 'Enviar':

            try:
                codigoPedido = int(valores['codigo_cliente'])
                precobb1 = valores['prc_bb_1kg']
                precobb2 = valores['prc_bb_2kg']
                precobb5 = valores['prc_bb_5kg']
                precogv1 = valores['prc_gv_1kg']
                precogv5 = valores['prc_gv_5kg']
                precoverm = valores['prc_verm']
                precopto = valores['prc_pto']
                precoscbb = valores['prc_sc_bb']
                precoscgv = valores['prc_sc_gv']
                qtdBB1Pedido = int(valores['qtd_bb_1kg'])
                qtdBB2Pedido = int(valores['qtd_bb_2kg'])
                qtdBB5Pedido = int(valores['qtd_bb_5kg'])
                qtdVermPedido = int(valores['qtd_verm'])
                qtdPretoPedido = int(valores['qtd_pto'])
                qtdGV1Pedido = int(valores['qtd_gv_1kg'])
                qtdGV5Pedido = int(valores['qtd_gv_5kg'])
                qtdSCBBPedido = int(valores['qtd_sc_bb'])
                qtdSCGVPedido = int(valores['qtd_sc_gv'])
                obsPedido = valores['obs']

                date = dataPedido
                codigo = codigoPedido
                cliente = clientes.itensArquivo(codigo)
                nome_cliente = cliente[0]
                cidade1 = cliente[1]
                pagamento = cliente[2]

                if valores['nota'] == True:
                    try:
                        pagamento.index('Ch')
                        notaPedido = 'S/N'
                        janela1.hide()
                        break
                    except:
                        print('Não é permitido enviar boleto sem nota')
                if valores['nota'] == False:
                    notaPedido = ''
                    janela1.hide()
                    break
            except:
                print('Digite um código válido')
    #Eventos Janela3
        # Função do Botão "OK"
        if janela == janela3 and evento == 'OK':
            janela3.hide()
            janela1.un_hide()
            dataPedido = valores['data_dia']
            janela1.Element('dataPrevia').update(value=f'Data: {dataPedido}')
        # Função do Botão "Voltar"
        if janela == janela3 and evento == 'Voltar':
            janela3.hide()
            janela1.un_hide()
    #Eventos Janela4
    #Abrir Bloco de Pedido no Excel
    diretorioexcel = os.getcwd()
    nomebloco = "BLOCOPROJETO.xlsx"
    print('CRIANDO ARQUIVO')
    arquivo = f"{diretorioexcel}\{nomebloco}"
    wb = load_workbook(arquivo)
    ws = wb.worksheets[0]

    #Programa Principal

    #Cabecalho
    ws['C2'] = f'{nome_cliente}'
    ws['C4'] = f'{cidade1}'
    ws['H6'] = f'{pagamento}'
    ws['Q2'] = f'{date}'
    ws['Q3'] = f'{notaPedido}'

    #Quantidades
    qtdbb1 = ws['A8'] = int(f'{qtdBB1Pedido}')
    qtdbb2 = ws['A9'] = int(f'{qtdBB2Pedido}')
    qtdbb5 = ws['A10'] = int(f'{qtdBB5Pedido}')
    qtdverm = ws['A11'] = int(f'{qtdVermPedido}')
    qtdpt = ws['A12'] = int(f'{qtdPretoPedido}')
    qtdgv1 = ws['A17'] = int(f'{qtdGV1Pedido}')
    qtdgv5 = ws['A18'] = int(f'{qtdGV5Pedido}')
    qtscbb = ws['A26'] = int(f'{qtdSCBBPedido}')
    qtscgv = ws['A27'] = int(f'{qtdSCGVPedido}')
    obser = ws['H33'] = f'{obsPedido}'

    #Preços Unitários
    if qtdBB1Pedido > 0:
        pubb1 = ws['O8'] = int(f'{precobb1}')
    if qtdBB2Pedido > 0:
        pubb2 = ws['O9'] = int(f'{precobb2}')
    if qtdBB5Pedido > 0:
        pubb5 = ws['O10'] = int(f'{precobb5}')
    if qtdVermPedido > 0:
        puv = ws['O11'] = int(f'{precoverm}')
    if qtdPretoPedido > 0:
        pup = ws['O12'] = int(f'{precopto}')
    if qtdGV1Pedido > 0:
        pugv1 = ws['O17'] = int(f'{precogv1}')
    if qtdGV5Pedido > 0:
        pugv5 = ws['O18'] = int(f'{precogv5}')
    if qtdSCBBPedido > 0:
        puscbb = ws['O26'] = int(f'{precoscbb}')
    if qtdSCGVPedido > 0:
        puscgv = ws['O27'] = int(f'{precoscgv}')

    #Salvar pedido XL
    diretorio = os.getcwd()

    pastaxl = 'PedidosXL'
    pastapdf = 'PedidosPDF'

    if os.path.isdir(pastaxl):  # vemos de este diretorio ja existe
        None
    else:
        os.mkdir(pastaxl)  # aqui criamos a pasta caso nao exista
        print('Pasta criada com sucesso!')

    if os.path.isdir(pastapdf):  # vemos se este diretorio ja existe
        None
    else:
        os.mkdir(pastapdf)  # aqui criamos a pasta caso nao exista
        print('Pasta criada com sucesso!')

    caminhoXL = f'{diretorio}\{pastaxl}'
    caminhoPDF = f'{diretorio}\{pastapdf}'

    nome_arquivo = fr'{caminhoXL}\Pedido_{nome_cliente.strip()}'

    try:
        wb.save(f'{nome_arquivo}.xlsx')
        print('ARQUIVO XL CRIADO')
        nome_arquivo_PDF = fr'{caminhoPDF}\Pedido_{nome_cliente.strip()}'

        input_file = fr'{nome_arquivo}.xlsx'
        # give your file name with valid path
        output_file = fr'{nome_arquivo_PDF}.pdf'
        # give valid output file name and path
        app = client.DispatchEx("Excel.Application")
        app.Interactive = False
        app.Visible = False
        Workbook = app.Workbooks.Open(input_file)
        try:
            Workbook.ActiveSheet.ExportAsFixedFormat(0, output_file)
            print('Arquivo PDF Criado!')
        except Exception as e:
            print(
                "Failed to convert in PDF format.Please confirm environment meets all the requirements  and try again")
            print(str(e))
        finally:
            try:
                Workbook.Close()
                app.Exit()
            except Exception as e:
                None

    except Exception as e:
        print(f'Já existe um arquivo com esse nome')
