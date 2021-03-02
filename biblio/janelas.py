def janelaCadastro():
    from PySimpleGUI import PySimpleGUI as sg
    from datetime import datetime
    from biblio import clientes
    from openpyxl import load_workbook
    from win32com import client
    from sys import exit
    import os

    data_atual = datetime.today()
    data_texto = data_atual.strftime('%d/%m/%y')
    dataPedido = data_texto
    date = data_texto

    precobb = 190
    precogv = 175
    precoverm = 205
    precopto = 220
    precoscbb = 365
    precoscgv = 340

    sg.theme('Reddit')
    layout = [
        [sg.Text('Cadastrar')],
        [sg.Text('Cliente'), sg.Input(key='clienteCadastro')],
        [sg.Text('Endereço'), sg.Input(key='endereçoCadastro')],
        [sg.Text('Cidade'), sg.Input(key='cidadeCadastro')],
        [sg.Text('Fone'), sg.Input(key='foneCadastro')],
        [sg.Text('CNPJ'), sg.Input(key='cnpjCadastro')],
        [sg.Text('Insc. Est.'), sg.Input(key='ieCadastro')],
        [sg.Text('Cond. de Pag.'), sg.Input(key='pagCadastro')],
        [sg.Text('Email'), sg.Input(key='emailCadastro')],
        [sg.Checkbox('S/N', key='nota'), sg.Button('Digitar Data'), sg.Button('Limpar Data')],
        [sg.Text('Quantidades'), sg.Text('Preços'), sg.Text('R$')],
        [sg.Text('Barbalho 1kg'), sg.Input(f'{int(0)}', key='qtd_bb_1kg', size=(10, 1)),
         sg.Button('+1', key='+bb1'),
         sg.Button('-1', key='-bb1'), sg.Input(f'{int(precobb)}', key='prc_bb_1kg', size=(10, 1)),
         sg.Button('+1', key='+bb1prc'), sg.Button('-1', key='-bb1prc')],
        [sg.Text('Barbalho 2kg'), sg.Input(f'{int(0)}', key='qtd_bb_2kg', size=(10, 1)),
         sg.Button('+1', key='+bb2'),
         sg.Button('-1', key='-bb2'), sg.Input(f'{int(precobb)}', key='prc_bb_2kg', size=(10, 1)),
         sg.Button('+1', key='+bb2prc'), sg.Button('-1', key='-bb2prc')],
        [sg.Text('Barbalho 5kg'), sg.Input(f'{int(0)}', key='qtd_bb_5kg', size=(10, 1)),
         sg.Button('+1', key='+bb5'),
         sg.Button('-1', key='-bb5'), sg.Input(f'{int(precobb)}', key='prc_bb_5kg', size=(10, 1)),
         sg.Button('+1', key='+bb5prc'), sg.Button('-1', key='-bb5prc')],
        [sg.Text('Vermelho     '), sg.Input(f'{int(0)}', key='qtd_verm', size=(10, 1)),
         sg.Button('+1', key='+verm'),
         sg.Button('-1', key='-verm'), sg.Input(f'{int(precoverm)}', key='prc_verm', size=(10, 1)),
         sg.Button('+1', key='+vermprc'), sg.Button('-1', key='-vermprc')],
        [sg.Text('Preto           '), sg.Input(f'{int(0)}', key='qtd_pto', size=(10, 1)),
         sg.Button('+1', key='+pto'),
         sg.Button('-1', key='-pto'), sg.Input(f'{int(precopto)}', key='prc_pto', size=(10, 1)),
         sg.Button('+1', key='+ptoprc'), sg.Button('-1', key='-ptoprc')],
        [sg.Text('Goval 1kg    '), sg.Input(f'{int(0)}', key='qtd_gv_1kg', size=(10, 1)),
         sg.Button('+1', key='+gv1'),
         sg.Button('-1', key='-gv1'), sg.Input(f'{int(precogv)}', key='prc_gv_1kg', size=(10, 1)),
         sg.Button('+1', key='+gv1prc'), sg.Button('-1', key='-gv1prc')],
        [sg.Text('Goval 5kg    '), sg.Input(f'{int(0)}', key='qtd_gv_5kg', size=(10, 1)),
         sg.Button('+1', key='+gv5'),
         sg.Button('-1', key='-gv5'), sg.Input(f'{int(precogv)}', key='prc_gv_5kg', size=(10, 1)),
         sg.Button('+1', key='+gv5prc'), sg.Button('-1', key='-gv5prc')],
        [sg.Text('Sc. Barbalho'), sg.Input(f'{int(0)}', key='qtd_sc_bb', size=(10, 1)),
         sg.Button('+1', key='+scbb'),
         sg.Button('-1', key='-scbb'), sg.Input(f'{int(precoscbb)}', key='prc_sc_bb', size=(10, 1)),
         sg.Button('+1', key='+scbbprc'), sg.Button('-1', key='-scbbprc')],
        [sg.Text('Sc. Goval     '), sg.Input(f'{int(0)}', key='qtd_sc_gv', size=(10, 1)),
         sg.Button('+1', key='+scgv'),
         sg.Button('-1', key='-scgv'), sg.Input(f'{int(precoscgv)}', key='prc_sc_gv', size=(10, 1)),
         sg.Button('+1', key='+scgvprc'), sg.Button('-1', key='-scgvprc')],
        [sg.Text('Observações'), sg.Input(key='obs', size=(42, 1))],
        [sg.Button('Enviar'), sg.Button('Fechar')],
    ]
    janela = sg.Window('janelaCadastro', layout=layout,finalize=True)

    def atualizarVisorSoma(valor, elemento):
        soma = int(valor) + int(1)
        try:
            janela.Element(elemento).update(value=soma)
        except Exception as e:
            print(f'erro {e}')

    def atualizarVisorSub(valor, elemento):
        if int(valor) == 0:
            sub = 0
        else:
            sub = int(valor) - int(1)
        try:
            janela.Element(elemento).update(value=sub)
        except Exception as e:
            print(f'erro {e}')

    while True:
        evento, valores = janela.read()

        if evento == 'Fechar':
            return janela.hide()


        if evento == '+bb1':
            atualizarVisorSoma(valores['qtd_bb_1kg'], 'qtd_bb_1kg')
        if evento == '-bb1':
            atualizarVisorSub(valores['qtd_bb_1kg'], 'qtd_bb_1kg')

        if evento == '+bb1prc':
            atualizarVisorSoma(valores['prc_bb_1kg'], 'prc_bb_1kg')
        if evento == '-bb1prc':
            atualizarVisorSub(valores['prc_bb_1kg'], 'prc_bb_1kg')

        if evento == '+bb2':
            atualizarVisorSoma(valores['qtd_bb_2kg'], 'qtd_bb_2kg')
        if evento == '-bb2':
            atualizarVisorSub(valores['qtd_bb_2kg'], 'qtd_bb_2kg')

        if evento == '+bb2prc':
            atualizarVisorSoma(valores['prc_bb_2kg'], 'prc_bb_2kg')
        if evento == '-bb2prc':
            atualizarVisorSub(valores['prc_bb_2kg'], 'prc_bb_2kg')

        if evento == '+bb5':
            atualizarVisorSoma(valores['qtd_bb_5kg'], 'qtd_bb_5kg')
        if evento == '-bb5':
            atualizarVisorSub(valores['qtd_bb_5kg'], 'qtd_bb_5kg')

        if evento == '+bb5prc':
            atualizarVisorSoma(valores['prc_bb_5kg'], 'prc_bb_5kg')
        if evento == '-bb5prc':
            atualizarVisorSub(valores['prc_bb_5kg'], 'prc_bb_5kg')

        if evento == '+verm':
            atualizarVisorSoma(valores['qtd_verm'], 'qtd_verm')
        if evento == '-verm':
            atualizarVisorSub(valores['qtd_verm'], 'qtd_verm')

        if evento == '+vermprc':
            atualizarVisorSoma(valores['prc_verm'], 'prc_verm')
        if evento == '-vermprc':
            atualizarVisorSub(valores['prc_verm'], 'prc_verm')

        if evento == '+pto':
            atualizarVisorSoma(valores['qtd_pto'], 'qtd_pto')
        if evento == '-pto':
            atualizarVisorSub(valores['qtd_pto'], 'qtd_pto')

        if evento == '+ptoprc':
            atualizarVisorSoma(valores['prc_pto'], 'prc_pto')
        if evento == '-ptoprc':
            atualizarVisorSub(valores['prc_pto'], 'prc_pto')

        if evento == '+gv1':
            atualizarVisorSoma(valores['qtd_gv_1kg'], 'qtd_gv_1kg')
        if evento == '-gv1':
            atualizarVisorSub(valores['qtd_gv_1kg'], 'qtd_gv_1kg')

        if evento == '+gv1prc':
            atualizarVisorSoma(valores['prc_gv_1kg'], 'prc_gv_1kg')
        if evento == '-gv1prc':
            atualizarVisorSub(valores['prc_gv_1kg'], 'prc_gv_1kg')

        if evento == '+gv5':
            atualizarVisorSoma(valores['qtd_gv_5kg'], 'qtd_gv_5kg')
        if evento == '-gv5':
            atualizarVisorSub(valores['qtd_gv_5kg'], 'qtd_gv_5kg')

        if evento == '+gv5prc':
            atualizarVisorSoma(valores['prc_gv_5kg'], 'prc_gv_5kg')
        if evento == '-gv5prc':
            atualizarVisorSub(valores['prc_gv_5kg'], 'prc_gv_5kg')

        if evento == '+scbb':
            atualizarVisorSoma(valores['qtd_sc_bb'], 'qtd_sc_bb')
        if evento == '-scbb':
            atualizarVisorSub(valores['qtd_sc_bb'], 'qtd_sc_bb')

        if evento == '+scbbprc':
            atualizarVisorSoma(valores['prc_sc_bb'], 'prc_sc_bb')
        if evento == '-scbbprc':
            atualizarVisorSub(valores['prc_sc_bb'], 'prc_sc_bb')

        if evento == '+scgv':
            atualizarVisorSoma(valores['qtd_sc_gv'], 'qtd_sc_gv')
        if evento == '-scgv':
            atualizarVisorSub(valores['qtd_sc_gv'], 'qtd_sc_gv')

        if evento == '+scgvprc':
            atualizarVisorSoma(valores['prc_sc_gv'], 'prc_sc_gv')
        if evento == '-scgvprc':
            atualizarVisorSub(valores['prc_sc_gv'], 'prc_sc_gv')

        if evento == 'Enviar':
            if valores['clienteCadastro'] and valores['endereçoCadastro'] and valores['cidadeCadastro'] and valores[
                'foneCadastro'] and valores['cnpjCadastro'] and valores['ieCadastro'] and valores['pagCadastro'] and \
                    valores['emailCadastro'] != '':
                # cabeçalho
                try:
                    nome_cliente = valores['clienteCadastro']
                    data = date
                    endereco = valores['endereçoCadastro']
                    cidade = valores['cidadeCadastro']
                    fone = valores['foneCadastro']
                    cnpj = valores['cnpjCadastro']
                    ie = valores['ieCadastro']
                    pag = valores['pagCadastro']
                    mail = valores['emailCadastro']

                    # preços
                    precobb1 = valores['prc_bb_1kg']
                    precobb2 = valores['prc_bb_2kg']
                    precobb5 = valores['prc_bb_5kg']
                    precogv1 = valores['prc_gv_1kg']
                    precogv5 = valores['prc_gv_5kg']
                    precoverm = valores['prc_verm']
                    precopto = valores['prc_pto']
                    precoscbb = valores['prc_sc_bb']
                    precoscgv = valores['prc_sc_gv']
                    # quantidades
                    qtdBB1Pedido = int(valores['qtd_bb_1kg'])
                    qtdBB2Pedido = int(valores['qtd_bb_2kg'])
                    qtdBB5Pedido = int(valores['qtd_bb_5kg'])
                    qtdVermPedido = int(valores['qtd_verm'])
                    qtdPretoPedido = int(valores['qtd_pto'])
                    qtdGV1Pedido = int(valores['qtd_gv_1kg'])
                    qtdGV5Pedido = int(valores['qtd_gv_5kg'])
                    qtdSCBBPedido = int(valores['qtd_sc_bb'])
                    qtdSCGVPedido = int(valores['qtd_sc_gv'])
                    # observção
                    obsPedido = valores['obs']
                    if valores['nota'] == True:
                        try:
                            pag.index('Ch')
                            notaPedido = 'S/N'
                            break
                        except:
                            print('Não é permitido enviar boleto sem nota')
                    else:
                        notaPedido = ''
                        break
                except Exception as e:
                    print(f'ERRO {e}')
            else:
                print('Preencha todos os campos')

    # Abrir Arquivo do bloco de pedido
    diretorioexcel = os.getcwd()
    nomebloco = "BLOCOPROJETO.xlsx"
    print('CRIANDO ARQUIVO')
    arquivo = f"{diretorioexcel}\{nomebloco}"
    wb = load_workbook(arquivo)
    ws = wb.worksheets[0]
    # Preencher células
    ws['C2'] = f'{nome_cliente}'
    ws['Q2'] = f'{date}'
    ws['D3'] = f'{endereco}'
    ws['C4'] = f'{cidade}'
    ws['P4'] = f'{fone}'
    ws['B5'] = f'{cnpj}'
    ws['N5'] = f'{ie}'
    ws['H6'] = f'{pag}'
    ws['L6'] = f'{mail}'
    ws['Q3'] = f'{notaPedido}'

    # Quantidades
    qtdbb1 = ws['A8'] = int(f'{qtdBB1Pedido}')
    qtdbb2 = ws['A9'] = int(f'{qtdBB2Pedido}')
    qtdbb5 = ws['A10'] = int(f'{qtdBB5Pedido}')
    qtdverm = ws['A11'] = int(f'{qtdVermPedido}')
    qtdpt = ws['A12'] = int(f'{qtdPretoPedido}')
    qtdgv1 = ws['A17'] = int(f'{qtdGV1Pedido}')
    qtdgv5 = ws['A18'] = int(f'{qtdGV5Pedido}')
    qtscbb = ws['A26'] = int(f'{qtdSCBBPedido}')
    qtscgv = ws['A27'] = int(f'{qtdSCGVPedido}')
    obser = ws['H33'] = f'{obsPedido}'
    # Preços Unitários
    if qtdBB1Pedido > 0:
        pubb1 = ws['O8'] = int(f'{precobb1}')
    if qtdBB2Pedido > 0:
        pubb2 = ws['O9'] = int(f'{precobb2}')
    if qtdBB5Pedido > 0:
        pubb5 = ws['O10'] = int(f'{precobb5}')
    if qtdVermPedido > 0:
        puv = ws['O11'] = int(f'{precoverm}')
    if qtdPretoPedido > 0:
        pup = ws['O12'] = int(f'{precopto}')
    if qtdGV1Pedido > 0:
        pugv1 = ws['O17'] = int(f'{precogv1}')
    if qtdGV5Pedido > 0:
        pugv5 = ws['O18'] = int(f'{precogv5}')
    if qtdSCBBPedido > 0:
        puscbb = ws['O26'] = int(f'{precoscbb}')
    if qtdSCGVPedido > 0:
        puscgv = ws['O27'] = int(f'{precoscgv}')

    # Salvar pedido XL
    diretorio = os.getcwd()

    pastaxl = 'CadastrosXL'
    pastapdf = 'CadastrosPDF'

    if os.path.isdir(pastaxl):  # vemos de este diretorio ja existe
        None
    else:
        os.mkdir(pastaxl)  # aqui criamos a pasta caso nao exista
        print('Pasta criada com sucesso!')

    if os.path.isdir(pastapdf):  # vemos se este diretorio ja existe
        None
    else:
        os.mkdir(pastapdf)  # aqui criamos a pasta caso nao exista
        print('Pasta criada com sucesso!')

    caminhoXL = f'{diretorio}\{pastaxl}'
    caminhoPDF = f'{diretorio}\{pastapdf}'

    nome_arquivo = fr'{caminhoXL}\Pedido_{nome_cliente.strip()}'

    try:
        wb.save(f'{nome_arquivo}.xlsx')
        print('ARQUIVO XL CRIADO')
        nome_arquivo_PDF = fr'{caminhoPDF}\Pedido_{nome_cliente.strip()}'

        input_file = fr'{nome_arquivo}.xlsx'
        # give your file name with valid path
        output_file = fr'{nome_arquivo_PDF}.pdf'
        # give valid output file name and path
        app = client.DispatchEx("Excel.Application")
        app.Interactive = False
        app.Visible = False
        Workbook = app.Workbooks.Open(input_file)
        try:
            Workbook.ActiveSheet.ExportAsFixedFormat(0, output_file)
            print('Arquivo PDF Criado!')
        except Exception as e:
            print(
                "Failed to convert in PDF format.Please confirm environment meets all the requirements  and try again")
            print(str(e))
        finally:
            try:
                Workbook.Close()
                app.Exit()

            except Exception as e:
                None

    except Exception as e:
        print(f'Já existe um arquivo com esse nome')

    arquivo = 'cadastrosclientes.txt'
    clientes.cadastrar(arquivo, nome_cliente, cidade, pag)
    janela.hide()
