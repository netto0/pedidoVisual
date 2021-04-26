from openpyxl import load_workbook
from biblio import clientes
from win32com import client
from sys import exit
from PySimpleGUI import PySimpleGUI as sg
from datetime import datetime
import time
import os

#preços
precoBB = 190
precoGV = 175
precoPTO = 220
precoVERM = 210
precoSCBB = 370
precoSCGV = 340
precoFARBB = 80
precoFARGV = 75
precoFARSC = 190


soma_qtd = 0
soma_sc = 0
qtd_bb_fd = 0
qtd_gv_fd = 0
total_bb_sc = 0
total_gv_sc = 0
total_far_bb = 0
total_far_gv = 0
qtd_far_sc = 0
total_verm = 0
total_pto = 0
total_far_sc = 0

caminho = f'{os.getcwd()}\{"PedidosXl"}'
arquivos = []
for item in os.listdir(caminho):
    arquivos.append(item)
cont = 0
for i in arquivos:
    try:
        arquivo = f'{os.getcwd()}\PedidosXl\{i}'
        wb = load_workbook(arquivo,data_only=True)
        ws = wb.worksheets[0]

        if ws['A8'].value != None: qtd_bb1 = ws['A8'].value
        else: qtd_bb1 = 0
        if ws['A9'].value != None:qtd_bb2 = ws['A9'].value
        else: qtd_bb2 = 0
        if ws['A10'].value != None:qtd_bb5 = ws['A10'].value
        else:
            qtd_bb5 = 0
        if ws['A11'].value != None:qtd_pto = ws['A12'].value
        else:
            qtd_pto = 0
        if ws['A12'].value != None:qtd_verm = ws['A11'].value
        else:
            qtd_verm = 0
        if ws['A17'].value != None:qtd_gv1 = ws['A17'].value
        else:
            qtd_gv1 = 0
        if ws['A18'].value != None:qtd_gv5 = ws['A18'].value
        else:
            qtd_gv5 = 0
        if ws['A23'].value != None:qtd_far_bb = ws['A23'].value
        else:
            qtd_far_bb = 0
        if ws['A24'].value != None:qtd_far_gv = ws['A24'].value
        else:
            qtd_far_gv = 0
        if ws['A26'].value != None:qtd_sc_bb = ws['A26'].value
        else:
            qtd_sc_bb = 0
        if ws['A27'].value != None:qtd_sc_gv = ws['A27'].value
        else:
            qtd_sc_gv = 0
        if ws['A28'].value != None:qtd_sc_far_bb = ws['A28'].value
        else:
            qtd_sc_far_bb = 0
        if ws['A29'].value != None:qtd_sc_far_gv = ws['A29'].value
        else:
            qtd_sc_far_gv = 0


        soma_qtd += int(qtd_bb1)
        soma_qtd += int(qtd_bb2)
        soma_qtd += int(qtd_bb5)
        qtd_bb_fd += int(qtd_bb1)
        qtd_bb_fd += int(qtd_bb2)
        qtd_bb_fd += int(qtd_bb5)

        soma_qtd += int(qtd_gv1)
        soma_qtd += int(qtd_gv5)
        qtd_gv_fd += int(qtd_gv1)
        qtd_gv_fd += int(qtd_gv5)

        soma_qtd += int(qtd_verm)
        total_verm += int(qtd_verm)

        soma_qtd += int(qtd_pto)
        total_pto += int(qtd_pto)

        soma_qtd += int(qtd_far_bb)
        total_far_bb += int(qtd_far_bb)

        soma_qtd += int(qtd_far_gv)
        total_far_gv += int(qtd_far_gv)

        soma_sc += int(qtd_sc_bb)
        total_bb_sc += int(qtd_sc_bb)

        soma_sc += int(qtd_sc_gv)
        total_gv_sc += int(qtd_sc_gv)

        soma_sc += int(qtd_sc_far_bb)
        total_far_sc += int(qtd_sc_far_bb)

        soma_sc += int(qtd_sc_far_gv)
        total_far_sc += int(qtd_sc_far_bb)

        cont += 1

    except Exception as e:
        print(e)

peso_fardo = soma_qtd * 30
peso_saco = soma_sc * 60
peso_Total = peso_fardo + peso_saco

valor_bb = qtd_bb_fd * precoBB
valor_gv = qtd_gv_fd * precoGV
valor_pto = total_pto * precoPTO
valor_verm = total_verm * precoVERM
valor_saco_bb = total_bb_sc * precoSCBB
valor_saco_gv = total_gv_sc * precoSCGV
valor_far_bb = total_far_bb * precoFARBB
valor_far_gv = total_far_gv * precoFARGV
valor_far_sc = total_far_sc * precoFARSC
valor_Total = valor_bb + valor_gv + valor_pto + valor_verm + valor_saco_bb + valor_saco_gv + valor_far_bb + valor_far_gv + valor_far_sc

print(f'Pedidos: {cont}'
      f'\n'
      f'Quantidade Barbalho: {qtd_bb_fd}\n'
      f'Quantidade Goval: {qtd_gv_fd}\n'
      f'Quantidade Preto: {total_pto}\n'
      f'Quantidade Vermelho: {total_verm}\n'
      f'Quantidade Saco Barbalho: {total_bb_sc}\n'
      f'Quantidade Saco Goval: {total_gv_sc}\n'
      f'\n'
      f'Peso Total: {peso_Total} Kgs\n'
      f'\n'
      f'Valor Barbalho: R$ {valor_bb}\n'
      f'Valor Goval: R$ {valor_gv}\n'
      f'Valor Preto: R$ {valor_pto}\n'
      f'Valor Vermelho: R$ {valor_verm}\n'
      f'Valor Saco Barbalho: R$ {valor_saco_bb}\n'
      f'Valor Saco Goval: R$ {valor_saco_gv}\n'
      f'Valor Farinha Barbalho: R$ {valor_far_bb}\n'
      f'Valor Farinha Goval: R$ {valor_far_gv}\n'
      f'Valor Farinha Saco: R$ {valor_far_sc}\n'
      f'Valor Total: R$ {valor_Total}')

